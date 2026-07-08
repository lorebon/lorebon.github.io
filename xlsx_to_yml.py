#!/usr/bin/env python3
"""
xlsx_to_yml.py  —  Convert publications.xlsx to publications.yml and
                   publications-list.qmd for a Quarto academic website.

The script keeps the original YAML output for Quarto/listing workflows, and
also writes a ready-to-include Markdown/Quarto fragment with this layout:

    authors
    title
    journal/proceedings/book venue (year)

Usage:
    python xlsx_to_yml.py
    python xlsx_to_yml.py input.xlsx output.yml
    python xlsx_to_yml.py input.xlsx output.yml output.qmd
    python xlsx_to_yml.py --force

Requirements:
    pip install openpyxl
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import openpyxl

# ── Default file paths ──────────────────────────────────────────────────────
DEFAULT_INPUT = "publications.xlsx"
DEFAULT_OUTPUT_YML = "publications.yml"
DEFAULT_OUTPUT_QMD = "publications-list.qmd"

# ── Column header → internal key mapping ────────────────────────────────────
# The extra keys make the script robust to journal articles, working papers,
# book chapters, and conference/proceedings entries.
KEY_MAP = {
    "Section": "section",
    "Authors": "authors",
    "Year": "year",
    "Date": "date",
    "Title": "title",
    "Paper Link": "path",
    "Link": "path",
    "URL": "path",
    "Journal": "journal",
    "Venue": "venue",
    "Book": "book",
    "Book Title": "booktitle",
    "Proceedings": "proceedings",
    "Conference": "conference",
    "Conference Proceedings": "conferenceproceedings",
    "Publisher": "publisher",
    "Editors": "editors",
    "Volume": "volume",
    "Issue": "issue",
    "Pages": "pages",
    "DOI": "doi",
    "PDF": "pdf",
    "Preprint": "preprint",
    "ShareIt": "shareit",
    "Supplemental Information": "supplemental",
    "GitHub": "github",
    "Code": "code",
    "Data": "data",
    "Highly Cited": "highlycited",
    "Hot Paper": "hotpaper",
    "Awards": "awards",
    "Media Coverage": "mediacoverage",
    "Invited Presentation": "invitedpresentation",
    "Categories": "categories",
}

INT_KEYS = {"year"}

SECTION_ALIASES = {
    "working paper": "Working Papers",
    "working papers": "Working Papers",
    "work in progress": "Working Papers",
    "preprint": "Working Papers",
    "preprints": "Working Papers",
    "peer-reviewed journal paper": "Journal Articles",
    "peer reviewed journal paper": "Journal Articles",
    "journal paper": "Journal Articles",
    "journal papers": "Journal Articles",
    "journal article": "Journal Articles",
    "journal articles": "Journal Articles",
    "article": "Journal Articles",
    "articles": "Journal Articles",
    "book chapter": "Book Chapters and Conference Proceedings",
    "book chapters": "Book Chapters and Conference Proceedings",
    "conference proceeding": "Book Chapters and Conference Proceedings",
    "conference proceedings": "Book Chapters and Conference Proceedings",
    "proceeding": "Book Chapters and Conference Proceedings",
    "proceedings": "Book Chapters and Conference Proceedings",
}

SECTION_ORDER = [
    "Working Papers",
    "Journal Articles",
    "Book Chapters and Conference Proceedings",
]

FIELD_ORDER = [
    "section",
    "authors",
    "authors_line",
    "year",
    "date",
    "title",
    "title_line",
    "path",
    "journal",
    "venue",
    "booktitle",
    "book",
    "proceedings",
    "conference",
    "conferenceproceedings",
    "publisher",
    "editors",
    "venue_line",
    "volume",
    "issue",
    "pages",
    "doi",
    "pdf",
    "preprint",
    "shareit",
    "supplemental",
    "github",
    "code",
    "data",
    "highlycited",
    "hotpaper",
    "awards",
    "mediacoverage",
    "invitedpresentation",
    "categories",
]


def yaml_scalar(value: Any) -> str:
    """Return a properly quoted YAML scalar string."""
    text = str(value).strip()
    if text == "":
        return '""'

    special = set(':#|>[]{}*!,%@`&')
    needs_quotes = (
        any(c in text for c in special)
        or text.startswith(('"', "'", "-", "?", "@", "`"))
        or text.lower() in {"true", "false", "null", "none", "yes", "no"}
        or "\n" in text
    )
    if needs_quotes:
        escaped = text.replace('\\', '\\\\').replace('"', '\\"')
        return f'"{escaped}"'
    return text


def yaml_inline_list(values: list[str]) -> str:
    """Return a YAML inline list: [a, b, "c d"]."""
    return "[" + ", ".join(yaml_scalar(v) for v in values) + "]"


def parse_categories(value: Any) -> list[str]:
    """Parse a categories cell into a list."""
    s = str(value).strip()
    if not s:
        return []
    for sep in (";", "|", ","):
        s = s.replace(sep, ",")
    return [item.strip() for item in s.split(",") if item.strip()]


def normalize_section(value: Any) -> str:
    """Map spreadsheet section labels to the three desired homepage groups."""
    raw = str(value).strip()
    if not raw:
        return "Working Papers"
    return SECTION_ALIASES.get(raw.lower(), raw)


def first_nonempty(rec: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        value = rec.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def make_primary_link(rec: dict[str, Any]) -> str:
    """
    Use only the Excel 'Paper Link' column for the clickable title.

    If Paper Link is empty, the title is rendered as plain text,
    even if DOI, PDF, or Preprint fields are present.
    """
    link = str(rec.get("path", "")).strip()
    if not link:
        return ""
    return link


def make_venue_line(rec: dict[str, Any]) -> str:
    """
    Build the third display line.

    Required form:
        Journal / venue / proceedings name (year)

    Volume, issue, and pages are appended after the year if available.
    """
    venue = first_nonempty(
        rec,
        [
            "journal",
            "venue",
            "conferenceproceedings",
            "proceedings",
            "booktitle",
            "book",
            "conference",
            "publisher",
        ],
    )

    year = str(rec.get("year", "")).strip()
    if venue and year:
        line = f"{venue} ({year})"
    elif venue:
        line = venue
    elif year:
        line = f"({year})"
    else:
        line = ""

    details = []
    volume = str(rec.get("volume", "")).strip()
    issue = str(rec.get("issue", "")).strip()
    pages = str(rec.get("pages", "")).strip()

    if volume and issue:
        details.append(f"{volume}({issue})")
    elif volume:
        details.append(volume)
    elif issue:
        details.append(f"({issue})")

    if pages:
        details.append(f"pp. {pages}")

    if line and details:
        line = f"{line}, " + ", ".join(details)
    elif details:
        line = ", ".join(details)

    return line


def enrich_record(rec: dict[str, Any]) -> dict[str, Any]:
    """Add normalized section and three display fields used by the homepage."""
    enriched = dict(rec)
    enriched["section"] = normalize_section(enriched.get("section", ""))
    enriched["authors_line"] = str(enriched.get("authors", "")).strip()
    enriched["title_line"] = str(enriched.get("title", "")).strip()
    venue_line = make_venue_line(enriched)
    if venue_line:
        enriched["venue_line"] = venue_line
    return enriched


def parse_row(headers: list[Any], row: tuple[Any, ...]) -> dict[str, Any]:
    """Convert one spreadsheet row into a dictionary of publication fields."""
    rec: dict[str, Any] = {}
    for header, cell_value in zip(headers, row):
        if header is None:
            continue
        header_text = str(header).strip()
        if not header_text or cell_value is None:
            continue

        key = KEY_MAP.get(header_text, header_text.lower().replace(" ", "-"))
        if key in INT_KEYS:
            try:
                rec[key] = int(cell_value)
            except (ValueError, TypeError):
                s = str(cell_value).strip()
                if s:
                    rec[key] = s
        elif key == "categories":
            cats = parse_categories(cell_value)
            if cats:
                rec[key] = cats
        else:
            s = str(cell_value).strip()
            if s:
                rec[key] = s
    return enrich_record(rec) if rec else rec


def record_to_yaml(rec: dict[str, Any]) -> str:
    """Serialize one publication record as a YAML list-item block."""
    lines = []
    first = True
    for key in FIELD_ORDER:
        if key not in rec:
            continue
        value = rec[key]
        prefix = "- " if first else "  "
        first = False

        if isinstance(value, int):
            lines.append(f"{prefix}{key}: {value}")
        elif isinstance(value, list) and key == "categories":
            cleaned = [str(v).strip() for v in value if str(v).strip()]
            if cleaned:
                lines.append(f"{prefix}{key}: {yaml_inline_list(cleaned)}")
        else:
            lines.append(f"{prefix}{key}: {yaml_scalar(value)}")

    return "\n".join(lines)


def sort_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    order = {section: i for i, section in enumerate(SECTION_ORDER)}

    def sort_key(rec: dict[str, Any]) -> tuple[int, int, str]:
        section_rank = order.get(str(rec.get("section", "")), len(order))
        try:
            year_rank = -int(rec.get("year", 0))
        except (ValueError, TypeError):
            year_rank = 0
        title_rank = str(rec.get("title", "")).lower()
        return (section_rank, year_rank, title_rank)

    return sorted(records, key=sort_key)


def markdown_escape(text: Any) -> str:
    """Escape the most common Markdown control characters in display text."""
    s = str(text).strip()
    for char in ["\\", "`", "*", "_", "{", "}", "[", "]", "<", ">", "#", "+", "-", ".", "!", "|"]:
        s = s.replace(char, "\\" + char)
    return s


def markdown_link(text: str, url: str) -> str:
    """Return a Markdown link, falling back to plain text if no URL is provided."""
    label = markdown_escape(text)
    link = str(url).strip()
    if not link:
        return label
    # Spaces in URLs are invalid in Markdown links; encode the common case.
    link = link.replace(" ", "%20")
    return f"[{label}]({link})"


def publication_entry_to_qmd(rec: dict[str, Any]) -> str:
    """Render one publication as exactly three visible Markdown lines."""
    authors = markdown_escape(rec.get("authors_line", ""))
    title = str(rec.get("title_line", "")).strip()
    venue = markdown_escape(rec.get("venue_line", ""))
    link = make_primary_link(rec)

    title_line = markdown_link(title, link) if title else ""

    lines = []
    if authors:
        lines.append(authors + "  ")
    if title_line:
        lines.append(f"**{title_line}**  ")
    if venue:
        lines.append(f"*{venue}*")

    return "\n".join(lines)


def records_to_qmd(records: list[dict[str, Any]]) -> str:
    """Render all publications as a Quarto fragment grouped by section."""
    grouped: dict[str, list[dict[str, Any]]] = {section: [] for section in SECTION_ORDER}
    extras: dict[str, list[dict[str, Any]]] = {}

    for rec in sort_records(records):
        section = str(rec.get("section", "")).strip() or "Working Papers"
        if section in grouped:
            grouped[section].append(rec)
        else:
            extras.setdefault(section, []).append(rec)

    chunks = [
        "<!-- Generated by xlsx_to_yml.py. Do not edit this file directly. -->",
        "",
    ]

    for section in SECTION_ORDER:
        section_records = grouped[section]
        if not section_records:
            continue
        chunks.append(f"### {section}")
        chunks.append("")
        for rec in section_records:
            chunks.append(publication_entry_to_qmd(rec))
            chunks.append("")

    for section, section_records in extras.items():
        chunks.append(f"### {section}")
        chunks.append("")
        for rec in section_records:
            chunks.append(publication_entry_to_qmd(rec))
            chunks.append("")

    if len(chunks) <= 2:
        chunks.append("No publications listed yet.")
        chunks.append("")

    return "\n".join(chunks).rstrip() + "\n"


def convert(input_path: str, output_yml_path: str, output_qmd_path: str) -> None:
    print(f"Reading  : {input_path}")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    print(f"Columns  : {headers}")

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rec = parse_row(headers, row)
        if rec:
            records.append(rec)

    records = sort_records(records)
    print(f"Records  : {len(records)}")

    yaml_text = "\n\n".join(record_to_yaml(r) for r in records) + "\n"
    Path(output_yml_path).write_text(yaml_text, encoding="utf-8")
    print(f"Written  : {output_yml_path}")

    qmd_text = records_to_qmd(records)
    Path(output_qmd_path).write_text(qmd_text, encoding="utf-8")
    print(f"Written  : {output_qmd_path}")

    try:
        import yaml  # type: ignore

        parsed = yaml.safe_load(Path(output_yml_path).read_text(encoding="utf-8"))
        parsed_count = len(parsed) if parsed else 0
        print(f"Validated: {parsed_count} YAML entries parsed successfully by PyYAML")
    except ImportError:
        print("Tip: install PyYAML (`pip install pyyaml`) for automatic validation")
    except Exception as exc:
        print(f"WARNING: YAML validation failed — {exc}")


def should_convert(input_path: str, outputs: list[str], force: bool = False) -> bool:
    """Return True when the input is newer than any output, or any output is missing."""
    if force:
        return True

    src = Path(input_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    for output in outputs:
        dst = Path(output)
        if not dst.exists():
            print(f"Output missing: {output}; running conversion.")
            return True
        if src.stat().st_mtime > dst.stat().st_mtime:
            print(f"Detected update in {input_path}; running conversion.")
            return True

    print(f"No update in {input_path}; skip conversion.")
    return False


if __name__ == "__main__":
    args = [arg for arg in sys.argv[1:] if arg != "--force"]
    force = "--force" in sys.argv[1:]

    input_file = args[0] if len(args) > 0 else DEFAULT_INPUT
    output_yml_file = args[1] if len(args) > 1 else DEFAULT_OUTPUT_YML
    output_qmd_file = args[2] if len(args) > 2 else DEFAULT_OUTPUT_QMD

    if should_convert(input_file, [output_yml_file, output_qmd_file], force=force):
        convert(input_file, output_yml_file, output_qmd_file)
